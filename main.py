import os
import sys
import json
import time
import base64
import logging
import requests
import mimetypes
import shutil
import re
import tempfile
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
import fitz  # PyMuPDF
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Alignment, numbers, PatternFill
from openpyxl.formatting.rule import CellIsRule
from typing import Any, Dict, List, Optional, Union
from requests.adapters import HTTPAdapter
from urllib3.util import Retry
from concurrent.futures import ThreadPoolExecutor, as_completed


# =============================================================================
# 유틸리티 함수들
# =============================================================================
def safe_json_dump(data: Any, filepath: Path) -> None:
    """JSON 데이터를 안전하게 파일에 저장"""
    try:
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logging.info(f"JSON 파일 저장 성공: {filepath}")
    except Exception as e:
        logging.exception(f"JSON 파일 저장 실패: {filepath}, 오류: {e}")


def safe_file_delete(file_path: Path) -> None:
    """파일 삭제를 시도하고 실패 시 예외를 로깅"""
    try:
        if file_path.exists():
            file_path.unlink()
    except Exception as e:
        logging.exception(f"파일 삭제 실패: {file_path}, 오류: {e}")


def extract_json_using_regex(content: str) -> str:
    """정규표현식을 사용하여 JSON 객체 부분만 추출"""
    pattern = r"({.*})"
    match = re.search(pattern, content, re.DOTALL)
    if match:
        return match.group(1)
    return content


def get_retry_session(
    retries: int = 3,
    backoff_factor: float = 1.0,
    status_forcelist: tuple = (500, 502, 503, 504),
    session: Optional[requests.Session] = None,
) -> requests.Session:
    """
    재시도 메커니즘이 포함된 requests 세션을 반환
    """
    session = session or requests.Session()
    retry = Retry(
        total=retries,
        read=retries,
        connect=retries,
        backoff_factor=backoff_factor,
        status_forcelist=status_forcelist,
        allowed_methods=frozenset(
            ["HEAD", "GET", "POST", "PUT", "DELETE", "OPTIONS", "TRACE"]
        ),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


# =============================================================================
# OCR 결과 처리 관련 함수
# =============================================================================
def ocr_result_to_jsonl_lines(ocr_result: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    OCR 결과 JSON을 줄 단위 JSONL 객체 리스트로 변환
    """
    lines: List[Dict[str, Any]] = []
    images = ocr_result.get("images", [])
    for image in images:
        fields = image.get("fields", [])
        current_line_text: str = ""
        confidences: List[float] = []
        line_number: int = 1
        for field in fields:
            text: str = field.get("inferText", "").strip()
            conf = field.get("inferConfidence", None)
            if conf is not None:
                confidences.append(conf)
            if text:
                current_line_text = (
                    f"{current_line_text} {text}".strip() if current_line_text else text
                )
            if field.get("lineBreak", False):
                if current_line_text:
                    avg_conf: Optional[float] = (
                        round(sum(confidences) / len(confidences), 2)
                        if confidences
                        else None
                    )
                    lines.append(
                        {
                            "line": line_number,
                            "text": current_line_text,
                            "신뢰도": avg_conf,
                        }
                    )
                    line_number += 1
                    current_line_text = ""
                    confidences = []
        if current_line_text:
            avg_conf = (
                round(sum(confidences) / len(confidences), 2) if confidences else None
            )
            lines.append(
                {
                    "line": line_number,
                    "text": current_line_text,
                    "신뢰도": avg_conf,
                }
            )
    return lines


# =============================================================================
# OpenAI API 분석 함수
# =============================================================================
def analyze_merged_jsonl(
    merged_text: str,
    openai_api_key: str,
    openai_api_url: str,
    session: Optional[requests.Session] = None,
    timeout: int = 60,
) -> Optional[Dict[str, Any]]:
    """
    병합된 JSONL 텍스트를 OpenAI API를 통해 분석하여 지정 항목을 추출
    """
    prompt = f"""다음은 병합된 jsonl 파일의 내용입니다. 각 줄은 OCR 결과의 일부입니다.
아래 항목들을 추출하여 JSON 형식으로 반환해줘.
항목:
- 접수일자: 문서에서 접수일자를 "YYYY-MM-DD" 형식으로 추출
- 세목: 문서에서 세목 정보를 추출하되, 민간경상보조, 민간위탁사업비, 사업출연금 3개 항목 중 1개로 선택
- 세부사업명: 문서에서 세부사업명을 추출하되, 확인되지 않으면 빈 항목으로 표시
- 내역사업명: 문서에서 사업명 정보를 추출 (내역사업명으로 표시)
- 기관명: 내역사업명 옆에 예금주와 동일한 값으로 표시
- 신뢰도: OCR 결과의 평균 신뢰도를 소수점 두 자리까지 표시 (Excel에서는 %로 표시)
- 연간집행계획액, 기수령액, 월집행계획액, 전월이월액, 당월신청액, 누계: 해당 금액은 jsonl 파일 전체의 총계이며, 숫자만 추출하고 천 단위마다 쉼표(,)를 추가하여 표시
- 은행명: 문서에서 은행명을 추출
- 계좌번호: 문서에서 계좌번호 정보를 추출
- 예금주: 문서에서 예금주 정보를 추출

병합된 jsonl 파일 내용:
{merged_text}
"""
    payload = {
        "model": "gpt-4o-mini",
        "messages": [
            {"role": "system", "content": "당신은 OCR 결과 분석 전문가입니다."},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.3,
        "max_tokens": 1000,
    }
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {openai_api_key}",
    }

    session = session or get_retry_session()
    try:
        response = session.post(
            openai_api_url, headers=headers, json=payload, timeout=timeout
        )
        if response.status_code == 200:
            result = response.json()
            choices = result.get("choices")
            if (
                not choices
                or not isinstance(choices, list)
                or "message" not in choices[0]
            ):
                logging.error("OpenAI API 응답 구조가 예상과 다릅니다.")
                return None
            content = choices[0]["message"].get("content", "")
            try:
                data = json.loads(content)
                return data
            except json.JSONDecodeError:
                cleaned = extract_json_using_regex(content)
                try:
                    data = json.loads(cleaned)
                    return data
                except json.JSONDecodeError:
                    logging.error("OpenAI 응답 JSON 파싱 실패")
                    logging.error(content)
                    return None
        else:
            logging.error(f"OpenAI API 호출 실패, 상태코드: {response.status_code}")
            logging.error(response.text)
            return None
    except requests.RequestException as e:
        logging.exception(f"OpenAI API 호출 중 오류 발생: {e}")
        return None


# =============================================================================
# 환경설정 및 디렉토리 생성
# =============================================================================
def setup_environment() -> Dict[str, Any]:
    """환경변수를 로드하고 검증하여 설정 딕셔너리를 반환"""
    load_dotenv(override=True)
    required_vars = [
        "CLOVA_OCR_APIGW_INVOKE_URL",
        "CLOVA_OCR_SECRET_KEY",
        "OPENAI_API_KEY",
    ]
    missing = [var for var in required_vars if not os.getenv(var)]
    if missing:
        logging.error(f"필수 환경변수가 누락되었습니다: {', '.join(missing)}")
        sys.exit(1)

    config = {
        "SOURCE_DIR": Path(os.getenv("SOURCE_DIR", "source")),
        "MERGED_DIR": Path(os.getenv("MERGED_DIR", "merged")),
        "RESULT_DIR": Path(os.getenv("RESULT_DIR", "result")),
        "LOG_DIR": Path(os.getenv("LOG_DIR", "logs")),
        "CLOVA_OCR_API_URL": os.getenv("CLOVA_OCR_APIGW_INVOKE_URL"),
        "CLOVA_OCR_SECRET": os.getenv("CLOVA_OCR_SECRET_KEY"),
        "MAX_FILE_SIZE": int(os.getenv("MAX_FILE_SIZE", 16777216)),
        "OPENAI_API_KEY": os.getenv("OPENAI_API_KEY"),
        "OPENAI_API_URL": "https://api.openai.com/v1/chat/completions",
    }
    return config


def setup_logging() -> None:
    """로그 설정 (로그 파일에만 기록)"""
    config = setup_environment()
    log_dir: Path = config["LOG_DIR"]
    log_dir.mkdir(parents=True, exist_ok=True)
    date_str: str = datetime.now().strftime("%Y%m%d")
    log_filename: Path = log_dir / f"log_{date_str}.log"
    file_handler = logging.FileHandler(log_filename, mode="a", encoding="utf-8")
    file_handler.setFormatter(
        logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    )
    logging.basicConfig(level=logging.INFO, handlers=[file_handler])
    logging.info(f"로그 파일 생성됨: {log_filename}")


def ensure_directories(dirs: List[Path]) -> None:
    """목록에 있는 모든 디렉토리를 생성"""
    for directory in dirs:
        directory.mkdir(parents=True, exist_ok=True)


# =============================================================================
# 파일 검증 및 OCR API 호출 관련 함수
# =============================================================================
VALID_EXTENSIONS = [".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".tif", ".bmp"]


def is_valid_file(file_path: Union[str, Path], max_size: int) -> bool:
    """파일 존재, 확장자, 크기, MIME 타입을 검사"""
    file_path = Path(file_path)
    if not file_path.exists():
        logging.error(f"파일이 존재하지 않습니다: {file_path}")
        return False
    if file_path.suffix.lower() not in VALID_EXTENSIONS:
        logging.error(f"지원되지 않는 파일 형식: {file_path.suffix.lower()}")
        return False
    if file_path.stat().st_size > max_size:
        logging.error(f"파일 크기가 너무 큽니다: {file_path.stat().st_size} 바이트")
        return False
    if file_path.suffix.lower() != ".pdf":
        mime_type, _ = mimetypes.guess_type(str(file_path))
        if not mime_type or not mime_type.startswith("image/"):
            logging.error(f"지원되지 않는 MIME 타입: {mime_type}")
            return False
    return True


def call_clova_ocr_api(
    file_path: Union[str, Path],
    api_url: str,
    secret: str,
    session: Optional[requests.Session] = None,
    timeout: int = 60,
) -> Optional[Dict[str, Any]]:
    """
    Clova OCR API를 호출하여 파일을 base64 인코딩 후 전송
    """
    file_path = Path(file_path)
    file_ext: str = file_path.suffix[1:].lower()
    headers = {"X-OCR-SECRET": secret, "Content-Type": "application/json"}
    try:
        with open(file_path, "rb") as f:
            file_data = f.read()
    except Exception as e:
        logging.exception(f"파일 읽기 실패: {file_path}, 오류: {e}")
        return None

    payload = {
        "version": "V2",
        "requestId": f"request_{int(time.time() * 1000)}",
        "timestamp": int(time.time() * 1000),
        "images": [
            {
                "format": file_ext.upper(),
                "name": file_path.name,
                "data": base64.b64encode(file_data).decode("utf-8"),
            }
        ],
    }

    session = session or get_retry_session()
    try:
        response = session.post(api_url, headers=headers, json=payload, timeout=timeout)
        if response.status_code == 200:
            logging.info(f"OCR API 호출 성공: {file_path}")
            return response.json()
        else:
            logging.error(
                f"OCR API 호출 실패: {file_path}, 상태코드: {response.status_code}"
            )
            return None
    except requests.RequestException as e:
        logging.exception(f"OCR API 호출 중 오류 발생: {file_path}, 오류: {e}")
        return None


# =============================================================================
# PDF를 이미지로 변환 (임시 디렉토리 사용)
# =============================================================================
def convert_pdf_to_images(
    pdf_path: Union[str, Path], temp_dir: Union[str, Path]
) -> List[Path]:
    """
    PDF 파일의 각 페이지를 이미지로 변환하여 임시 디렉토리에 저장한 후 이미지 경로 리스트 반환
    """
    pdf_path = Path(pdf_path)
    temp_dir = Path(temp_dir)
    image_paths: List[Path] = []
    try:
        with fitz.open(pdf_path) as pdf:
            for page_num in range(len(pdf)):
                page = pdf.load_page(page_num)
                matrix = fitz.Matrix(300 / 72, 300 / 72)
                pix = page.get_pixmap(matrix=matrix)
                image_path: Path = temp_dir / f"page_{page_num+1}.png"
                pix.save(str(image_path))
                image_paths.append(image_path)
        logging.info(f"PDF 변환 완료: {pdf_path} -> {len(image_paths)} 이미지")
    except Exception as e:
        logging.exception(f"PDF 변환 오류: {pdf_path}, 오류: {e}")
    return image_paths


# =============================================================================
# 파일별 OCR 처리 및 JSONL, 원본 OCR 결과 파일 생성
# =============================================================================
def process_file_to_jsonl(
    file_path: Union[str, Path],
    config: Dict[str, Any],
    target_folder: Path,
    session: Optional[requests.Session] = None,
) -> Optional[Path]:
    """
    파일별로 OCR을 수행하여 원본 OCR 결과와 JSONL 파일을 생성
    """
    file_path = Path(file_path)
    if not is_valid_file(file_path, config["MAX_FILE_SIZE"]):
        return None

    ocr_results: List[Dict[str, Any]] = []
    # PDF 파일인 경우, 임시 디렉토리를 사용하여 이미지 변환
    if file_path.suffix.lower() == ".pdf":
        with tempfile.TemporaryDirectory() as tmpdirname:
            temp_dir = Path(tmpdirname)
            image_paths = convert_pdf_to_images(file_path, temp_dir)
            if not image_paths:
                logging.error(f"PDF 변환 실패: {file_path}")
                return None
            for img_path in image_paths:
                result = call_clova_ocr_api(
                    img_path,
                    config["CLOVA_OCR_API_URL"],
                    config["CLOVA_OCR_SECRET"],
                    session=session,
                )
                if result:
                    ocr_results.append(result)
                safe_file_delete(img_path)
    else:
        result = call_clova_ocr_api(
            file_path,
            config["CLOVA_OCR_API_URL"],
            config["CLOVA_OCR_SECRET"],
            session=session,
        )
        if result:
            ocr_results.append(result)

    target_folder.mkdir(parents=True, exist_ok=True)
    raw_ocr_path: Path = target_folder / f"{file_path.stem}.ocr.json"
    safe_json_dump(ocr_results, raw_ocr_path)

    all_lines: List[Dict[str, Any]] = []
    for ocr_result in ocr_results:
        lines = ocr_result_to_jsonl_lines(ocr_result)
        for obj in lines:
            obj["source_file"] = file_path.name
        all_lines.extend(lines)
    jsonl_path: Path = target_folder / f"{file_path.stem}.jsonl"
    try:
        with open(jsonl_path, "w", encoding="utf-8") as f:
            for obj in all_lines:
                f.write(json.dumps(obj, ensure_ascii=False) + "\n")
        logging.info(f"JSONL 파일 저장: {jsonl_path}")
        return jsonl_path
    except Exception as e:
        logging.exception(f"JSONL 파일 저장 실패: {file_path}, 오류: {e}")
        return None


# =============================================================================
# 하위 폴더 내 JSONL 파일들을 병합
# =============================================================================
def process_subfolder_target(target_folder: Union[str, Path]) -> None:
    """
    대상 폴더 내의 JSONL 파일들을 하나의 병합 파일로 생성
    """
    target_folder = Path(target_folder)
    jsonl_files: List[Path] = list(target_folder.glob("*.jsonl"))
    if not jsonl_files:
        logging.info(f"대상 폴더에 JSONL 파일이 없습니다: {target_folder}")
        return
    jsonl_files.sort(key=lambda f: f.name)
    merged_lines: List[str] = []
    for jf in jsonl_files:
        try:
            with open(jf, "r", encoding="utf-8") as infile:
                merged_lines.extend([line.strip() for line in infile if line.strip()])
        except Exception as e:
            logging.exception(f"JSONL 파일 읽기 실패: {jf}, 오류: {e}")
    merged_file_path: Path = target_folder / f"{target_folder.name}.jsonl"
    try:
        with open(merged_file_path, "w", encoding="utf-8") as outfile:
            for line in merged_lines:
                outfile.write(line + "\n")
        logging.info(f"병합된 JSONL 파일 생성: {merged_file_path}")
    except Exception as e:
        logging.exception(f"병합 실패: {target_folder}, 오류: {e}")


# =============================================================================
# 전체 작업을 단계별로 실행 (파일 처리, JSONL 병합, OpenAI 분석)
# =============================================================================
def run_all_tasks(
    config: Dict[str, Any], session: requests.Session
) -> List[Dict[str, Any]]:
    """
    전체 작업을 단계별로 실행:
     - 파일 처리 (OCR 및 JSONL 생성)
     - JSONL 병합
     - OpenAI 분석
    각 단계는 별도의 진행률바로 표시됨.
    """
    overall_results = []
    source_dir: Path = config["SOURCE_DIR"]
    merged_dir: Path = config["MERGED_DIR"]

    # 1. 파일 처리 작업 (OCR 및 JSONL 생성)
    file_tasks = []
    for subfolder in source_dir.iterdir():
        if subfolder.is_dir():
            target_folder: Path = merged_dir / subfolder.name
            target_folder.mkdir(parents=True, exist_ok=True)
            files = [
                file
                for file in subfolder.iterdir()
                if file.is_file() and file.suffix.lower() in VALID_EXTENSIONS
            ]
            for file in files:
                file_tasks.append((file, target_folder))

    with tqdm(
        total=len(file_tasks),
        desc="파일 처리 (OCR 및 JSONL 생성)",
        ncols=100,
        colour="green",
    ) as pbar_file:
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = [
                executor.submit(
                    process_file_to_jsonl, file, config, target_folder, session
                )
                for file, target_folder in file_tasks
            ]
            for _ in as_completed(futures):
                pbar_file.update(1)

    # 2. JSONL 병합 작업
    merge_tasks = []
    for subfolder in merged_dir.iterdir():
        if subfolder.is_dir() and list(subfolder.glob("*.jsonl")):
            merge_tasks.append(subfolder)

    with tqdm(
        total=len(merge_tasks), desc="JSONL 병합", ncols=100, colour="yellow"
    ) as pbar_merge:
        for subfolder in merge_tasks:
            process_subfolder_target(subfolder)
            pbar_merge.update(1)

    # 3. OpenAI 분석 작업
    analysis_tasks = []
    for subfolder in merged_dir.iterdir():
        if subfolder.is_dir():
            merged_file = subfolder / f"{subfolder.name}.jsonl"
            if merged_file.exists():
                analysis_tasks.append(merged_file)

    with tqdm(
        total=len(analysis_tasks), desc="OpenAI 분석", ncols=100, colour="blue"
    ) as pbar_analysis:
        for merged_file in analysis_tasks:
            try:
                with open(merged_file, "r", encoding="utf-8") as f:
                    merged_text = f.read()
                logging.info(f"분석 시작: {merged_file.name}")
                result = analyze_merged_jsonl(
                    merged_text,
                    config["OPENAI_API_KEY"],
                    config["OPENAI_API_URL"],
                    session=session,
                    timeout=60,
                )
                if result:
                    overall_results.append(result)
            except Exception as e:
                logging.exception(f"분석 파일 처리 실패: {merged_file}, 오류: {e}")
            pbar_analysis.update(1)

    return overall_results


# =============================================================================
# Excel 파일 내보내기
# =============================================================================
def export_excel_from_data(
    data_list: List[Dict[str, Any]], output_path: Union[str, Path]
) -> None:
    """
    분석된 데이터를 Excel 파일로 내보내기
    (신뢰도는 소수점 형태로 저장 후 "0.00%" 서식을 적용)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "결과"
    headers = [
        "접수일자",
        "세목",
        "세부사업명",
        "내역사업명",
        "기관명",
        "신뢰도",
        "연간집행계획액",
        "기수령액",
        "월집행계획액",
        "전월이월액",
        "당월신청액",
        "누계",
        "은행명",
        "계좌번호",
        "예금주",
    ]
    ws.append(headers)
    amount_fields = {
        "연간집행계획액",
        "기수령액",
        "월집행계획액",
        "전월이월액",
        "당월신청액",
        "누계",
    }

    for data in data_list:
        row = []
        for col in headers:
            if col == "기관명":
                value = data.get("예금주", "")
            elif col == "신뢰도":
                try:
                    numeric_val = float(data.get("신뢰도", 0))
                    value = numeric_val
                except (ValueError, TypeError):
                    value = data.get("신뢰도", "")
            elif col in amount_fields:
                try:
                    numeric_value = int(float(data.get(col, 0)))
                    value = f"{numeric_value:,}"
                except (ValueError, TypeError):
                    value = str(data.get(col, ""))
            else:
                value = data.get(col, "")
            row.append(value)
        ws.append(row)

    for col_idx, header in enumerate(headers, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if header in amount_fields:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            if header == "신뢰도" and cell.row != 1:
                cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="center")

    max_row = ws.max_row
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )

    ws.conditional_formatting.add(
        f"F2:F{max_row}",
        CellIsRule(
            operator="lessThan", formula=["0.5"], stopIfTrue=True, fill=red_fill
        ),
    )
    ws.conditional_formatting.add(
        f"F2:F{max_row}",
        CellIsRule(
            operator="between",
            formula=["0.5", "0.8"],
            stopIfTrue=True,
            fill=yellow_fill,
        ),
    )
    ws.conditional_formatting.add(
        f"F2:F{max_row}",
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["0.8"],
            stopIfTrue=True,
            fill=green_fill,
        ),
    )

    wb.save(output_path)
    logging.info(f"Excel 파일 저장 완료: {output_path}")


# =============================================================================
# 메인 함수
# =============================================================================
def main() -> None:
    """전체 작업(파일 처리, JSONL 병합, OpenAI 분석, Excel 내보내기)을 단계별 진행률바로 실행"""
    config = setup_environment()
    setup_logging()
    ensure_directories(
        [
            config["SOURCE_DIR"],
            config["MERGED_DIR"],
            config["RESULT_DIR"],
            config["LOG_DIR"],
        ]
    )

    # 모든 API 호출에 재사용할 세션 생성
    session = get_retry_session()

    # 단계별 진행: 파일 처리, JSONL 병합, OpenAI 분석
    results = run_all_tasks(config, session)

    # Excel 파일 생성 단계 (진행률바 1단계)
    output_excel = (
        config["RESULT_DIR"] / f"output_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )
    with tqdm(
        total=1, desc="Excel 파일 생성", ncols=100, colour="magenta"
    ) as pbar_excel:
        export_excel_from_data(results, output_excel)
        pbar_excel.update(1)

    logging.info(
        "모든 작업(파일 처리, JSONL 병합, OpenAI 분석, Excel 내보내기) 완료되었습니다."
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logging.exception(f"메인 실행 중 예상치 못한 오류 발생: {e}")
        sys.exit(1)
